# -*- coding: utf-8 -*-
"""Cálculo_da_relação_Economias_População.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1ZU7uREOC-Csu3OceqmERAoRSK3jAX9oM
"""

import openpyxl

wb = openpyxl.load_workbook('base.xlsx')
sh1 = wb['BASE']
for i in range(802,5,-5):
  sh1.insert_rows(i,2)
  for j in range(1,4):
    sh1.cell(i,j).value = sh1.cell(i-1,j).value
    sh1.cell(i+1,j).value = sh1.cell(i-1,j).value
  sh1.cell(i,4).value = 'PopÁgua/EcoSAA'
  sh1.cell(i+1,4).value = 'PopEsgoto/EcoSES'
  sh1.cell(i,5).value = 'hab/n'
  sh1.cell(i+1,5).value = 'hab/n'
  for j in range(6,36):
    try:
      sh1.cell(i,j).value = float(sh1.cell(i-4,j).value) / float(sh1.cell(i-2,j).value)
    except:
      sh1.cell(i,j).value = 'NDA'
    try:
      sh1.cell(i+1,j).value = float(sh1.cell(i-3,j).value) / float(sh1.cell(i-1,j).value)
    except:
      sh1.cell(i+1,j).value = 'NDA'
wb.save('base2.xlsx')